import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ─── Color Palette ───────────────────────────────────────────────
DARK_BG     = "1E1E2E"   # deep navy/charcoal
ACCENT      = "7C3AED"   # violet
ACCENT2     = "A78BFA"   # lavender
INCOME_GRN  = "059669"   # emerald
INCOME_LT   = "D1FAE5"   # mint tint
EXPENSE_RED = "DC2626"   # red
EXPENSE_LT  = "FEE2E2"   # rose tint
HEADER_BG   = "1E1E2E"
HEADER_FG   = "FFFFFF"
ROW_ALT     = "F5F3FF"   # very light purple
ROW_WHITE   = "FFFFFF"
BORDER_CLR  = "DDD6FE"
GOLD        = "F59E0B"
DASHBOARD   = "0F172A"   # darkest navy

thin = Side(style="thin", color=BORDER_CLR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_style(cell, bg=HEADER_BG, fg=HEADER_FG, size=11, bold=True, center=True):
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=True)
    cell.border = border

def data_style(cell, bg=ROW_WHITE, number_format=None, bold=False, fg="1E1E2E", center=False):
    cell.font = Font(color=fg, size=10, name="Calibri", bold=bold)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center")
    cell.border = border
    if number_format:
        cell.number_format = number_format

def currency(cell, bg=ROW_WHITE, fg="1E1E2E", bold=False):
    data_style(cell, bg=bg, number_format='"$"#,##0.00', fg=fg, bold=bold, center=True)

def pct(cell, bg=ROW_WHITE):
    cell.number_format = "0.0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ════════════════════════════════════════════════════════════════
#  WORKBOOK
# ════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ────────────────────────────────────────────────────────────────
#  SHEET 1: DASHBOARD
# ────────────────────────────────────────────────────────────────
ws_dash = wb.active
ws_dash.title = "📊 Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.row_dimensions[1].height = 8

# Title banner
ws_dash.merge_cells("B2:K4")
title_cell = ws_dash["B2"]
title_cell.value = "SMALL BUSINESS BOOKKEEPING"
title_cell.font = Font(bold=True, color="FFFFFF", size=22, name="Calibri")
title_cell.fill = PatternFill("solid", fgColor=DASHBOARD)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws_dash.merge_cells("B5:K5")
sub = ws_dash["B5"]
sub.value = "Income · Expenses · Profit & Loss  |  2025"
sub.font = Font(bold=False, color=ACCENT2, size=11, name="Calibri")
sub.fill = PatternFill("solid", fgColor=DASHBOARD)
sub.alignment = Alignment(horizontal="center", vertical="center")

ws_dash.row_dimensions[2].height = 36
ws_dash.row_dimensions[5].height = 20

# KPI Cards row
kpi_labels = ["TOTAL INCOME", "TOTAL EXPENSES", "NET PROFIT", "PROFIT MARGIN"]
kpi_cells  = ["C7", "E7", "G7", "I7"]
val_cells  = ["C8", "E8", "G8", "I8"]
kpi_colors = [INCOME_GRN, EXPENSE_RED, ACCENT, GOLD]

for i, (lbl, lc, vc, clr) in enumerate(zip(kpi_labels, kpi_cells, val_cells, kpi_colors)):
    col = ord(lc[0]) - 64
    ws_dash.merge_cells(f"{get_column_letter(col)}7:{get_column_letter(col+1)}7")
    ws_dash.merge_cells(f"{get_column_letter(col)}8:{get_column_letter(col+1)}8")
    c_lbl = ws_dash[lc]
    c_lbl.value = lbl
    c_lbl.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    c_lbl.fill = PatternFill("solid", fgColor=clr)
    c_lbl.alignment = Alignment(horizontal="center", vertical="center")
    c_val = ws_dash[vc]
    if i < 3:
        c_val.value = f"=SUM('💰 Income'!G2:G1000)" if i == 0 else (
                       f"=SUM('💸 Expenses'!G2:G1000)" if i == 1 else
                       f"=SUM('💰 Income'!G2:G1000)-SUM('💸 Expenses'!G2:G1000)")
        c_val.number_format = '"$"#,##0.00'
    else:
        c_val.value = f"=IF(SUM('💰 Income'!G2:G1000)=0,0,(SUM('💰 Income'!G2:G1000)-SUM('💸 Expenses'!G2:G1000))/SUM('💰 Income'!G2:G1000))"
        c_val.number_format = "0.0%"
    c_val.font = Font(bold=True, color=clr, size=16, name="Calibri")
    c_val.fill = PatternFill("solid", fgColor="FAFAF9")
    c_val.alignment = Alignment(horizontal="center", vertical="center")
    c_val.border = Border(left=Side(style="medium",color=clr),
                          right=Side(style="medium",color=clr),
                          bottom=Side(style="medium",color=clr))

ws_dash.row_dimensions[7].height = 22
ws_dash.row_dimensions[8].height = 34

# Monthly summary header
ws_dash.merge_cells("B10:K10")
mh = ws_dash["B10"]
mh.value = "MONTHLY SUMMARY"
mh.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
mh.fill = PatternFill("solid", fgColor=ACCENT)
mh.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[10].height = 22

months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
month_headers = ["MONTH","INCOME","EXPENSES","NET","MARGIN"]
col_map = [2,3,5,7,9]
bg_heads = [DARK_BG, INCOME_GRN, EXPENSE_RED, ACCENT, GOLD]

for ci, (hdr, col, bg) in enumerate(zip(month_headers, col_map, bg_heads)):
    c = ws_dash.cell(row=11, column=col, value=hdr)
    header_style(c, bg=bg)
    if ci > 0:
        ws_dash.merge_cells(start_row=11, start_column=col, end_row=11, end_column=col+1)

ws_dash.row_dimensions[11].height = 20

for ri, mon in enumerate(months):
    row = 12 + ri
    bg = ROW_ALT if ri % 2 == 0 else ROW_WHITE
    # Month label
    c_mon = ws_dash.cell(row=row, column=2, value=mon)
    data_style(c_mon, bg=bg, bold=True, center=True)
    # Income
    inc_f = f"=SUMPRODUCT((MONTH('💰 Income'!A2:A1000)={ri+1})*('💰 Income'!G2:G1000))"
    exp_f = f"=SUMPRODUCT((MONTH('💸 Expenses'!A2:A1000)={ri+1})*('💸 Expenses'!G2:G1000))"
    c_inc = ws_dash.cell(row=row, column=3, value=inc_f)
    ws_dash.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    currency(c_inc, bg=INCOME_LT, fg=INCOME_GRN)
    c_exp = ws_dash.cell(row=row, column=5, value=exp_f)
    ws_dash.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    currency(c_exp, bg=EXPENSE_LT, fg=EXPENSE_RED)
    net_f = f"={get_column_letter(3)}{row}-{get_column_letter(5)}{row}"
    c_net = ws_dash.cell(row=row, column=7, value=net_f)
    ws_dash.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    c_net.number_format = '"$"#,##0.00'
    c_net.font = Font(bold=True, color=ACCENT, size=10, name="Calibri")
    c_net.fill = PatternFill("solid", fgColor=bg)
    c_net.alignment = Alignment(horizontal="center", vertical="center")
    c_net.border = border
    mg_f = f"=IF({get_column_letter(3)}{row}=0,0,({get_column_letter(7)}{row})/{get_column_letter(3)}{row})"
    c_mg = ws_dash.cell(row=row, column=9, value=mg_f)
    ws_dash.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
    c_mg.number_format = "0.0%"
    c_mg.font = Font(color=GOLD, size=10, name="Calibri", bold=True)
    c_mg.fill = PatternFill("solid", fgColor=bg)
    c_mg.alignment = Alignment(horizontal="center", vertical="center")
    c_mg.border = border
    ws_dash.row_dimensions[row].height = 18

# Totals row
tot_row = 12 + 12
ws_dash.merge_cells(f"B{tot_row}:B{tot_row}")
c_tot = ws_dash.cell(row=tot_row, column=2, value="TOTAL")
header_style(c_tot, bg=DARK_BG)
for col in [3, 5, 7, 9]:
    c = ws_dash.cell(row=tot_row, column=col)
    ws_dash.merge_cells(start_row=tot_row, start_column=col, end_row=tot_row, end_column=col+1)
    if col == 3:
        c.value = f"=SUM('💰 Income'!G2:G1000)"
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.fill = PatternFill("solid", fgColor=INCOME_GRN)
    elif col == 5:
        c.value = f"=SUM('💸 Expenses'!G2:G1000)"
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.fill = PatternFill("solid", fgColor=EXPENSE_RED)
    elif col == 7:
        c.value = f"=C{tot_row}-E{tot_row}"
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.fill = PatternFill("solid", fgColor=ACCENT)
    elif col == 9:
        c.value = f"=IF(C{tot_row}=0,0,G{tot_row}/C{tot_row})"
        c.number_format = "0.0%"
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws_dash.row_dimensions[tot_row].height = 22

# Column widths for dashboard
for c, w in [(2,12),(3,14),(4,2),(5,14),(6,2),(7,14),(8,2),(9,12),(10,2),(11,4)]:
    set_col_width(ws_dash, c, w)

# ────────────────────────────────────────────────────────────────
#  SHEET 2: INCOME
# ────────────────────────────────────────────────────────────────
ws_inc = wb.create_sheet("💰 Income")
ws_inc.sheet_view.showGridLines = False
ws_inc.freeze_panes = "A2"

income_headers = [
    "DATE", "CLIENT / CUSTOMER", "DESCRIPTION", "CATEGORY",
    "INVOICE #", "PAYMENT METHOD", "AMOUNT", "STATUS", "NOTES"
]
income_widths = [14, 26, 30, 20, 14, 18, 14, 14, 26]

# Header row
for col, (hdr, w) in enumerate(zip(income_headers, income_widths), start=1):
    c = ws_inc.cell(row=1, column=col, value=hdr)
    header_style(c, bg=INCOME_GRN)
    set_col_width(ws_inc, col, w)
ws_inc.row_dimensions[1].height = 22

# Sample data
income_data = [
    ["2025-01-05", "Acme Corp",       "Web design project",       "Design",      "INV-001", "Bank Transfer", 2500.00, "Paid",    ""],
    ["2025-01-18", "Jane Smith",      "Monthly retainer",         "Consulting",  "INV-002", "PayPal",         800.00, "Paid",    ""],
    ["2025-02-03", "Green Solutions", "Logo package",             "Design",      "INV-003", "Credit Card",    650.00, "Paid",    ""],
    ["2025-02-20", "TechStart Inc.",  "SEO audit",                "Marketing",   "INV-004", "Bank Transfer", 1200.00, "Paid",    ""],
    ["2025-03-10", "Local Bakery",    "Brand identity",           "Design",      "INV-005", "Check",         1800.00, "Pending", "Follow-up needed"],
    ["2025-03-28", "Riverside LLC",   "Social media management",  "Marketing",   "INV-006", "Bank Transfer",  950.00, "Paid",    ""],
    ["2025-04-07", "Global Events",   "Event photography",        "Photography", "INV-007", "PayPal",        1350.00, "Paid",    ""],
    ["2025-04-22", "Acme Corp",       "Website maintenance",      "Development", "INV-008", "Bank Transfer",  600.00, "Paid",    "Monthly"],
]

categories_inc = ["Design","Consulting","Marketing","Photography","Development","Coaching","Writing","Other"]
statuses = ["Paid","Pending","Overdue","Cancelled"]
methods = ["Bank Transfer","PayPal","Credit Card","Cash","Check","Venmo","Zelle","Stripe"]

# Dropdowns for income
dv_cat = DataValidation(type="list", formula1=f'"{",".join(categories_inc)}"', allow_blank=True)
dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses)}"', allow_blank=True)
dv_method = DataValidation(type="list", formula1=f'"{",".join(methods)}"', allow_blank=True)
dv_cat.sqref = "D2:D1000"
dv_status.sqref = "H2:H1000"
dv_method.sqref = "F2:F1000"
ws_inc.add_data_validation(dv_cat)
ws_inc.add_data_validation(dv_status)
ws_inc.add_data_validation(dv_method)

for ri, row_data in enumerate(income_data, start=2):
    bg = INCOME_LT if ri % 2 == 0 else ROW_WHITE
    for ci, val in enumerate(row_data, start=1):
        c = ws_inc.cell(row=ri, column=ci, value=val)
        if ci == 7:
            currency(c, bg=bg, fg=INCOME_GRN, bold=True)
        elif ci == 1:
            c.number_format = "YYYY-MM-DD"
            data_style(c, bg=bg, center=True)
        elif ci == 8:
            # Color status
            if val == "Paid":
                data_style(c, bg="D1FAE5", fg=INCOME_GRN, bold=True, center=True)
            elif val == "Pending":
                data_style(c, bg="FEF3C7", fg="D97706", bold=True, center=True)
            elif val == "Overdue":
                data_style(c, bg=EXPENSE_LT, fg=EXPENSE_RED, bold=True, center=True)
            else:
                data_style(c, bg=bg, center=True)
        else:
            data_style(c, bg=bg)
    ws_inc.row_dimensions[ri].height = 18

# Totals
tot_r = len(income_data) + 2
ws_inc.merge_cells(f"E{tot_r}:F{tot_r}")
ct = ws_inc.cell(row=tot_r, column=5, value="TOTAL INCOME")
header_style(ct, bg=INCOME_GRN, center=False)
cv = ws_inc.cell(row=tot_r, column=7, value=f"=SUM(G2:G{tot_r-1})")
currency(cv, bg=INCOME_LT, fg=INCOME_GRN, bold=True)
ws_inc.row_dimensions[tot_r].height = 22

# ────────────────────────────────────────────────────────────────
#  SHEET 3: EXPENSES
# ────────────────────────────────────────────────────────────────
ws_exp = wb.create_sheet("💸 Expenses")
ws_exp.sheet_view.showGridLines = False
ws_exp.freeze_panes = "A2"

expense_headers = [
    "DATE", "VENDOR / PAYEE", "DESCRIPTION", "CATEGORY",
    "PAYMENT METHOD", "RECEIPT #", "AMOUNT", "TAX DEDUCTIBLE?", "NOTES"
]
expense_widths = [14, 26, 30, 20, 18, 14, 14, 16, 26]

for col, (hdr, w) in enumerate(zip(expense_headers, expense_widths), start=1):
    c = ws_exp.cell(row=1, column=col, value=hdr)
    header_style(c, bg=EXPENSE_RED)
    set_col_width(ws_exp, col, w)
ws_exp.row_dimensions[1].height = 22

expense_data = [
    ["2025-01-03", "Adobe Creative",  "Monthly subscription",    "Software",      "Credit Card",   "REC-001", 54.99,  "Yes", ""],
    ["2025-01-10", "Office Depot",    "Desk supplies",           "Office",        "Credit Card",   "REC-002", 87.32,  "Yes", ""],
    ["2025-01-15", "Zoom",            "Pro plan",                "Software",      "Credit Card",   "REC-003", 14.99,  "Yes", ""],
    ["2025-02-01", "Shopify",         "Monthly plan",            "Software",      "Bank Transfer", "REC-004", 39.00,  "Yes", ""],
    ["2025-02-14", "Starbucks",       "Client meeting",          "Meals",         "Credit Card",   "REC-005", 23.75,  "Yes", "50% deductible"],
    ["2025-03-05", "AT&T",            "Business phone",          "Utilities",     "Bank Transfer", "REC-006", 85.00,  "Yes", ""],
    ["2025-03-20", "Canva Pro",       "Annual subscription",     "Software",      "Credit Card",   "REC-007", 119.99, "Yes", "Annual"],
    ["2025-04-02", "Amazon",          "Camera equipment",        "Equipment",     "Credit Card",   "REC-008", 349.00, "Yes", ""],
    ["2025-04-18", "Mailchimp",       "Email marketing",         "Marketing",     "Credit Card",   "REC-009", 30.00,  "Yes", ""],
]

categories_exp = ["Software","Office","Meals","Utilities","Equipment","Marketing","Travel","Contractor","Insurance","Rent","Other"]
yesno = ["Yes","No","Partial"]

dv_cat_e = DataValidation(type="list", formula1=f'"{",".join(categories_exp)}"', allow_blank=True)
dv_mth_e = DataValidation(type="list", formula1=f'"{",".join(methods)}"', allow_blank=True)
dv_tax   = DataValidation(type="list", formula1=f'"{",".join(yesno)}"', allow_blank=True)
dv_cat_e.sqref = "D2:D1000"
dv_mth_e.sqref = "E2:E1000"
dv_tax.sqref   = "H2:H1000"
ws_exp.add_data_validation(dv_cat_e)
ws_exp.add_data_validation(dv_mth_e)
ws_exp.add_data_validation(dv_tax)

for ri, row_data in enumerate(expense_data, start=2):
    bg = EXPENSE_LT if ri % 2 == 0 else ROW_WHITE
    for ci, val in enumerate(row_data, start=1):
        c = ws_exp.cell(row=ri, column=ci, value=val)
        if ci == 7:
            currency(c, bg=bg, fg=EXPENSE_RED, bold=True)
        elif ci == 1:
            c.number_format = "YYYY-MM-DD"
            data_style(c, bg=bg, center=True)
        elif ci == 8:
            if val == "Yes":
                data_style(c, bg="D1FAE5", fg=INCOME_GRN, bold=True, center=True)
            elif val == "No":
                data_style(c, bg=EXPENSE_LT, fg=EXPENSE_RED, bold=True, center=True)
            else:
                data_style(c, bg="FEF3C7", fg="D97706", bold=True, center=True)
        else:
            data_style(c, bg=bg)
    ws_exp.row_dimensions[ri].height = 18

tot_r_e = len(expense_data) + 2
ws_exp.merge_cells(f"E{tot_r_e}:F{tot_r_e}")
ct_e = ws_exp.cell(row=tot_r_e, column=5, value="TOTAL EXPENSES")
header_style(ct_e, bg=EXPENSE_RED, center=False)
cv_e = ws_exp.cell(row=tot_r_e, column=7, value=f"=SUM(G2:G{tot_r_e-1})")
currency(cv_e, bg=EXPENSE_LT, fg=EXPENSE_RED, bold=True)
ws_exp.row_dimensions[tot_r_e].height = 22

# ────────────────────────────────────────────────────────────────
#  SHEET 4: PROFIT & LOSS
# ────────────────────────────────────────────────────────────────
ws_pl = wb.create_sheet("📈 Profit & Loss")
ws_pl.sheet_view.showGridLines = False

def pl_section_header(ws, row, text, bg):
    ws.merge_cells(f"B{row}:H{row}")
    c = ws.cell(row=row, column=2, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24

def pl_row(ws, row, label, formula_or_val, is_total=False, indent=2):
    bg = ROW_WHITE if not is_total else "EDE9FE"
    ws.merge_cells(f"B{row}:E{row}")
    cl = ws.cell(row=row, column=2, value=label)
    cl.font = Font(bold=is_total, color=DARK_BG, size=10 if not is_total else 11, name="Calibri",
                   italic=not is_total and indent > 1)
    cl.fill = PatternFill("solid", fgColor=bg)
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    cl.border = border

    ws.merge_cells(f"F{row}:H{row}")
    cv = ws.cell(row=row, column=6, value=formula_or_val)
    cv.number_format = '"$"#,##0.00'
    cv.font = Font(bold=is_total, color=ACCENT if is_total else DARK_BG,
                   size=11 if is_total else 10, name="Calibri")
    cv.fill = PatternFill("solid", fgColor=bg)
    cv.alignment = Alignment(horizontal="right", vertical="center")
    cv.border = border
    ws.row_dimensions[row].height = 20
    return cv

# Header
ws_pl.merge_cells("B2:H4")
ph = ws_pl["B2"]
ph.value = "PROFIT & LOSS STATEMENT"
ph.font = Font(bold=True, color="FFFFFF", size=20, name="Calibri")
ph.fill = PatternFill("solid", fgColor=DASHBOARD)
ph.alignment = Alignment(horizontal="center", vertical="center")
ws_pl.row_dimensions[2].height = 40

ws_pl.merge_cells("B5:H5")
ps = ws_pl["B5"]
ps.value = "Fiscal Year 2025"
ps.font = Font(bold=False, color=ACCENT2, size=11, name="Calibri")
ps.fill = PatternFill("solid", fgColor=DASHBOARD)
ps.alignment = Alignment(horizontal="center", vertical="center")
ws_pl.row_dimensions[5].height = 20

r = 7
pl_section_header(ws_pl, r, "REVENUE", INCOME_GRN); r += 1

income_cats_pl = ["Design","Consulting","Marketing","Photography","Development","Coaching","Writing","Other"]
for cat in income_cats_pl:
    f = f"=SUMIF('💰 Income'!D:D,\"{cat}\",'💰 Income'!G:G)"
    pl_row(ws_pl, r, cat, f); r += 1

pl_row(ws_pl, r, "TOTAL REVENUE",
       f"=SUMIF('💰 Income'!H:H,\"Paid\",'💰 Income'!G:G)",
       is_total=True); r += 2

pl_section_header(ws_pl, r, "EXPENSES", EXPENSE_RED); r += 1

expense_cats_pl = ["Software","Office","Meals","Utilities","Equipment","Marketing","Travel","Contractor","Insurance","Rent","Other"]
for cat in expense_cats_pl:
    f = f"=SUMIF('💸 Expenses'!D:D,\"{cat}\",'💸 Expenses'!G:G)"
    pl_row(ws_pl, r, cat, f); r += 1

pl_row(ws_pl, r, "TOTAL EXPENSES",
       "=SUM('💸 Expenses'!G:G)",
       is_total=True); r += 2

pl_section_header(ws_pl, r, "NET PROFIT / LOSS", ACCENT); r += 1

net_r = r
pl_row(ws_pl, r, "Net Profit / Loss",
       "=SUM('💰 Income'!G:G)-SUM('💸 Expenses'!G:G)",
       is_total=True, indent=1)
r += 1
pl_row(ws_pl, r, "Profit Margin",
       "=IF(SUM('💰 Income'!G:G)=0,0,(SUM('💰 Income'!G:G)-SUM('💸 Expenses'!G:G))/SUM('💰 Income'!G:G))",
       is_total=True, indent=1)
ws_pl[f"F{r}"].number_format = "0.0%"

for col, w in [(2,30),(3,10),(4,10),(5,10),(6,12),(7,12),(8,5)]:
    set_col_width(ws_pl, col, w)

# ────────────────────────────────────────────────────────────────
#  SHEET 5: CATEGORY TRACKER
# ────────────────────────────────────────────────────────────────
ws_cat = wb.create_sheet("🏷️ Categories")
ws_cat.sheet_view.showGridLines = False

ws_cat.merge_cells("B2:I3")
ch = ws_cat["B2"]
ch.value = "SPENDING BY CATEGORY"
ch.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
ch.fill = PatternFill("solid", fgColor=DARK_BG)
ch.alignment = Alignment(horizontal="center", vertical="center")
ws_cat.row_dimensions[2].height = 30
ws_cat.row_dimensions[3].height = 10

cat_headers = ["CATEGORY", "INCOME", "EXPENSES", "NET", "% OF INCOME", "% OF EXPENSES"]
cat_col_w   = [22, 16, 16, 16, 14, 14]

for ci, (hdr, w) in enumerate(zip(cat_headers, cat_col_w), start=2):
    c = ws_cat.cell(row=4, column=ci, value=hdr)
    header_style(c, bg=DARK_BG)
    set_col_width(ws_cat, ci, w)
ws_cat.row_dimensions[4].height = 22

all_cats = sorted(set(categories_inc + categories_exp))
for ri, cat in enumerate(all_cats, start=5):
    bg = ROW_ALT if ri % 2 == 0 else ROW_WHITE
    c_cat = ws_cat.cell(row=ri, column=2, value=cat)
    data_style(c_cat, bg=bg, bold=True)
    inc_f = f"=SUMIF('💰 Income'!D:D,\"{cat}\",'💰 Income'!G:G)"
    exp_f = f"=SUMIF('💸 Expenses'!D:D,\"{cat}\",'💸 Expenses'!G:G)"
    c_inc = ws_cat.cell(row=ri, column=3, value=inc_f)
    currency(c_inc, bg=INCOME_LT, fg=INCOME_GRN)
    c_exp = ws_cat.cell(row=ri, column=4, value=exp_f)
    currency(c_exp, bg=EXPENSE_LT, fg=EXPENSE_RED)
    c_net = ws_cat.cell(row=ri, column=5, value=f"=C{ri}-D{ri}")
    c_net.number_format = '"$"#,##0.00'
    c_net.font = Font(bold=True, color=ACCENT, size=10, name="Calibri")
    c_net.fill = PatternFill("solid", fgColor=bg)
    c_net.alignment = Alignment(horizontal="center", vertical="center")
    c_net.border = border
    pct_inc = ws_cat.cell(row=ri, column=6,
                           value=f"=IF(SUM('💰 Income'!G:G)=0,0,C{ri}/SUM('💰 Income'!G:G))")
    pct_inc.number_format = "0.0%"
    pct_inc.font = Font(color=INCOME_GRN, size=10, name="Calibri")
    pct_inc.fill = PatternFill("solid", fgColor=bg)
    pct_inc.alignment = Alignment(horizontal="center", vertical="center")
    pct_inc.border = border
    pct_exp = ws_cat.cell(row=ri, column=7,
                           value=f"=IF(SUM('💸 Expenses'!G:G)=0,0,D{ri}/SUM('💸 Expenses'!G:G))")
    pct_exp.number_format = "0.0%"
    pct_exp.font = Font(color=EXPENSE_RED, size=10, name="Calibri")
    pct_exp.fill = PatternFill("solid", fgColor=bg)
    pct_exp.alignment = Alignment(horizontal="center", vertical="center")
    pct_exp.border = border
    ws_cat.row_dimensions[ri].height = 18

# ────────────────────────────────────────────────────────────────
#  SHEET ORDER & SHEET TABS
# ────────────────────────────────────────────────────────────────
tab_colors = {
    "📊 Dashboard":    "7C3AED",
    "💰 Income":       "059669",
    "💸 Expenses":     "DC2626",
    "📈 Profit & Loss":"F59E0B",
    "🏷️ Categories":   "0EA5E9",
}
for ws in wb.worksheets:
    if ws.title in tab_colors:
        ws.sheet_properties.tabColor = tab_colors[ws.title]

# ────────────────────────────────────────────────────────────────
#  SAVE
# ────────────────────────────────────────────────────────────────
out = os.path.expanduser("~/Downloads/SmallBusiness_Bookkeeping_2025.xlsx")
wb.save(out)
print(f"Saved: {out}")
